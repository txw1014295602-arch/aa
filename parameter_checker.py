#!/usr/bin/env python3
"""
参数核查系统 - 重新设计版本
支持双分表结构、复杂条件表达式、嵌套验证规则
Parameter Checker System - Redesigned Version
Supports dual-table structure, complex condition expressions, nested validation rules
"""

import pandas as pd
import logging
import re
from typing import Dict, List, Any, Optional, Set, Tuple, Callable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('parameter_checker.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)


def get_group_keys(df):
    return set(df.groupby(['f_site_id', 'f_cell_id']).groups.keys())
class ParameterChecker:
    """
    参数核查器类 - 支持双分表结构和复杂嵌套验证
    """
    # 类级常量：支持的运算符（按长度排序，确保长运算符优先匹配）
    OPERATORS = ['=','>=', '<=', '!=', '>', '<']
    # 逻辑运算符正则（用于匹配完整单词）
    LOGICAL_OP_PATTERN = re.compile(r'\band\b|\bor\b', re.IGNORECASE)
    def __init__(self, knowledge_file="参数知识库.xlsx"):
        """初始化参数核查器"""
        self.knowledge_file = knowledge_file
        self.parameter_info: Dict[str, Dict[str, Any]] = {}  # 参数信息表
        self.validation_rules: Dict[str, Dict[str, Any]] = {}  # 验证规则表

        # 加载知识库
        self.load_knowledge_base(knowledge_file)

    def load_knowledge_base(self, file_path: str) -> bool:
        """加载双分表知识库"""
        try:
            # 加载参数信息表
            param_success = self.load_parameter_info(file_path, "参数信息")
            # 加载验证规则表
            rule_success = self.load_validation_rules(file_path, "验证规则")

            if param_success and rule_success:
                logger.info(f"知识库加载成功: {len(self.parameter_info)}个参数, {len(self.validation_rules)}个验证规则")
                return True
            else:
                logger.warning("知识库加载部分失败")
                return False

        except FileNotFoundError:
            logger.info(f"文件 {file_path} 不存在，正在生成示例文件...")
            # 生成示例文件
            self.create_sample_excel()
            # 重新加载
            try:
                param_success = self.load_parameter_info(file_path, "参数信息")
                rule_success = self.load_validation_rules(file_path, "验证规则")
                if param_success and rule_success:
                    logger.info(
                        f"示例知识库加载成功: {len(self.parameter_info)}个参数, {len(self.validation_rules)}个验证规则")
                    return True
            except Exception as e:
                logger.error(f"重新加载知识库失败: {str(e)}")
            return False
        except Exception as e:
            logger.error(f"加载知识库失败: {str(e)}")
            return False

    def load_parameter_info(self, file_path: str, sheet_name: str) -> bool:
        """加载参数信息表"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

            # 验证必要列
            required_columns = ['MO名称', 'MO描述', '场景类型', '参数名称', '参数ID', '参数类型', '参数含义', '值描述']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                logger.error(f"参数信息表缺少必要列: {missing_columns}")
                return False

            self.parameter_info = {}

            # 按MO名称和参数名称分组
            for (mo_name, param_name), group in df.groupby(['MO名称', '参数名称'], dropna=False):
                # 使用第一行的信息
                row = group.iloc[0]

                # 初始化MO信息
                if mo_name not in self.parameter_info:
                    self.parameter_info[mo_name] = {
                        'mo_description': str(row.get('MO描述', '')).strip(),
                        'scenario': str(row.get('场景类型', '')).strip(),
                        'parameters': {}
                    }

                # 添加参数信息
                param_type = str(row.get('参数类型', 'single')).strip()
                self.parameter_info[mo_name]['parameters'][param_name] = {
                    'parameter_id': str(row.get('参数ID', '')).strip(),
                    'parameter_type': param_type,
                    'parameter_description': str(row.get('参数含义', '')).strip(),
                    'value_description': str(row.get('值描述', '')).strip()
                }

            logger.info(f"参数信息表加载成功: {len(self.parameter_info)} 个MO")
            return True

        except FileNotFoundError:
            raise  # 重新抛出FileNotFoundError让上层处理
        except Exception as e:
            logger.error(f"加载参数信息表失败: {str(e)}")
            return False

    def load_validation_rules(self, file_path: str, sheet_name: str) -> bool:
        """加载验证规则表"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

            # 验证必要列
            required_columns = ['校验ID', '校验类型', 'MO名称', '条件表达式', '期望值表达式', '错误描述', '继续校验ID']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                logger.error(f"验证规则表缺少必要列: {missing_columns}")
                return False

            self.validation_rules = {}

            for _, row in df.iterrows():
                rule_id = str(row.get('校验ID', '')).strip()
                if not rule_id or rule_id == 'nan':
                    continue

                self.validation_rules[rule_id] = {
                    'rule_id': rule_id,
                    'check_type': str(row.get('校验类型', '')).strip(),
                    'mo_name': str(row.get('MO名称', '')).strip(),
                    'condition_expression': str(row.get('条件表达式', '')).strip(),
                    'expected_expression': str(row.get('期望值表达式', '')).strip(),
                    'error_description': str(row.get('错误描述', '')).strip(),
                    'next_check_id': str(row.get('继续校验ID', '')).strip() if str(
                        row.get('继续校验ID', '')).strip() != 'nan' else None
                }

            logger.info(f"验证规则表加载成功: {len(self.validation_rules)} 个规则")
            return True

        except FileNotFoundError:
            raise  # 重新抛出FileNotFoundError让上层处理
        except Exception as e:
            logger.error(f"加载验证规则表失败: {str(e)}")
            return False

    def parse_condition_expression(self, expression: str, data_row: Dict[str, Any]) -> bool:
        """
        解析复杂条件表达式
        支持格式: （参数名1=值1and参数名2=值2）or（参数名3>值3and参数名2!=值2）
        """
        if not expression or expression == 'nan':
            return True

        try:
            # 标准化表达式：处理中文括号和操作符
            expression = self._normalize_condition_expression(expression)

            # 解析单个条件
            def evaluate_single_condition(cond: str) -> bool:
                cond = cond.strip()

                # 支持的运算符（按长度降序排列避免匹配问题）
                operators = ['>=', '<=', '!=', '>', '<', '=']

                for op in operators:
                    if op in cond:
                        parts = cond.split(op, 1)
                        if len(parts) == 2:
                            param_name = parts[0].strip()
                            expected_value = parts[1].strip()

                            # 获取实际值
                            actual_value = str(data_row.get(param_name, '')).strip()

                            # 执行比较
                            return self._compare_values(actual_value, op, expected_value)

                logger.warning(f"无法解析条件: {cond}")
                return False

            # 递归处理括号和逻辑运算符
            def evaluate_expression(expr: str) -> bool:
                expr = expr.strip()

                # 处理最内层括号
                while '(' in expr and ')' in expr:
                    start = expr.rfind('(')
                    end = expr.find(')', start)
                    if end == -1:
                        break

                    # 评估括号内的表达式
                    inner_expr = expr[start + 1:end]
                    result = self._evaluate_simple_expression(inner_expr, evaluate_single_condition)

                    # 替换括号及其内容为结果
                    expr = expr[:start] + str(result).lower() + expr[end + 1:]

                # 评估剩余表达式
                return self._evaluate_simple_expression(expr, evaluate_single_condition)

            return evaluate_expression(expression)

        except Exception as e:
            logger.error(f"解析条件表达式失败: {expression}, 错误: {str(e)}")
            return False

    def _normalize_condition_expression(self, expression: str) -> str:
        """标准化条件表达式：处理中文括号和操作符，表达式正则，and or前后加空格，' and '"""
        import re

        # 将中文括号转换为英文括号
        expression = expression.replace('（', '(').replace('）', ')')

        # 在逻辑操作符前后添加空格，处理无空格的情况
        expression = re.sub(r'(?<!\s)and(?!\s)', ' and ', expression)
        expression = re.sub(r'(?<!\s)or(?!\s)', ' or ', expression)

        # 处理中文逻辑操作符
        expression = expression.replace('且', ' and ').replace('或', ' or ')

        # 清理多余空格
        expression = re.sub(r'\s+', ' ', expression)

        return expression.strip()

    def _evaluate_simple_expression(self, expr: str, eval_func) -> bool:
        """评估简单表达式（不含括号）"""
        expr = expr.strip()

        # 处理 or 运算符
        if ' or ' in expr:
            parts = expr.split(' or ')
            return any(self._evaluate_simple_expression(part.strip(), eval_func) for part in parts)

        # 处理 and 运算符
        if ' and ' in expr:
            parts = expr.split(' and ')
            return all(self._evaluate_simple_expression(part.strip(), eval_func) for part in parts)

        # 处理布尔值
        if expr.lower() == 'true':
            return True
        elif expr.lower() == 'false':
            return False

        # 单个条件
        return eval_func(expr)

    def _compare_values(self, actual: str, operator: str, expected: str) -> bool:
        """比较两个值"""
        try:
            # 尝试数值比较
            try:
                actual_num = float(actual)
                expected_num = float(expected)

                if operator == '=':
                    return actual_num == expected_num
                elif operator == '!=':
                    return actual_num != expected_num
                elif operator == '>':
                    return actual_num > expected_num
                elif operator == '<':
                    return actual_num < expected_num
                elif operator == '>=':
                    return actual_num >= expected_num
                elif operator == '<=':
                    return actual_num <= expected_num
            except ValueError:
                # 字符串比较
                if operator == '=':
                    return actual == expected
                elif operator == '!=':
                    return actual != expected
                else:
                    # 字符串不支持大小比较
                    logger.warning(f"字符串不支持运算符 {operator}: {actual} {operator} {expected}")
                    return False

        except Exception as e:
            logger.error(f"值比较失败: {actual} {operator} {expected}, 错误: {str(e)}")
            return False

        return False


    def _is_complex_expression(self, expression: str) -> bool:
        """判断是否为复杂表达式（包含括号或逻辑运算符）"""
        has_parentheses = '(' in expression and ')' in expression
        has_logical_ops = self.LOGICAL_OP_PATTERN.search(expression) is not None
        return has_parentheses or has_logical_ops

    def _parse_complex_expression(self, expression: str) -> Dict[str, Any]:
        """解析复杂表达式"""
        try:
            processed_expr = self._process_parentheses(expression)
            params_info = self._extract_param_details(processed_expr)

            return {
                'type': 'complex',
                'expression': expression,
                'params': params_info
            }
        except Exception as e:
            # 复杂表达式解析失败时返回原始表达式和错误信息
            return {
                'type': 'complex',
                'expression': expression,
                'params': [],
                'error': f"解析复杂表达式失败: {str(e)}"
            }

    def _process_parentheses(self, expression: str) -> str:
        """处理表达式中的括号内容（支持多层嵌套）"""
        # 使用正则找到所有括号对
        paren_pattern = re.compile(r'\(([^()]*)\)')
        temp_expr = expression

        # 循环处理所有括号，直到没有括号为止
        while '(' in temp_expr and ')' in temp_expr:
            match = paren_pattern.search(temp_expr)
            if not match:
                break

            # 提取括号内内容并转为小写
            inner_expr = match.group(1).strip().lower()
            # 替换原括号部分
            temp_expr = f"{temp_expr[:match.start()]}{inner_expr}{temp_expr[match.end():]}"

        return temp_expr

    def _extract_param_details(self, expression: str) -> List[Dict[str, Any]]:
        """
        从表达式中提取参数详细信息
        统一处理单值/多值参数，返回结构化信息
        """
        params_info = []
        working_expr = expression.strip()

        while working_expr:
            found = False
            for op in self.OPERATORS:
                op_index = working_expr.find(op)
                if op_index != -1:
                    # 提取参数名
                    param_name = working_expr[:op_index].strip()
                    if not param_name:
                        working_expr = working_expr[op_index + len(op):].strip()
                        continue

                    # 提取参数值和更新工作表达式
                    param_value, working_expr = self._extract_param_value(
                        working_expr, op_index, len(op))

                    # 解析参数详情（单值/多值）
                    param_detail = self._parse_param_detail(
                        param_name, param_value, op)
                    params_info.append(param_detail)

                    found = True
                    break

            if not found:
                break

        return params_info

    def _extract_param_value(self, expr: str, op_index: int, op_length: int) -> Tuple[str, str]:
        """提取参数值并计算下一个处理位置"""
        remaining_expr = expr[op_index + op_length:].strip()

        # 查找下一个逻辑运算符
        and_match = self.LOGICAL_OP_PATTERN.search(remaining_expr, pos=0)
        or_match = self.LOGICAL_OP_PATTERN.search(remaining_expr, pos=0)

        # 计算值的结束位置
        delimiters = []
        if and_match:
            delimiters.append(and_match.start())
        if or_match:
            delimiters.append(or_match.start())

        if delimiters:
            value_end = min(delimiters)
            param_value = remaining_expr[:value_end].strip()
            # 计算下一个参数的起始位置（跳过逻辑运算符）
            next_op = and_match if and_match and and_match.start() == value_end else or_match
            next_start = value_end + len(next_op.group()) if next_op else len(remaining_expr)
            next_working_expr = remaining_expr[next_start:].strip()
        else:
            param_value = remaining_expr.strip()
            next_working_expr = ""

        return param_value, next_working_expr

    def _parse_simple_expression(self, expression: str) -> Dict[str, Any]:
        """解析简单表达式（逗号分隔的参数列表）"""
        expected_params = []
        # 按逗号分割多个参数表达式
        param_expressions = [expr.strip() for expr in expression.split(',') if expr.strip()]

        for param_expr in param_expressions:
            param_detail = self._parse_single_param_expr(param_expr)
            if param_detail:
                expected_params.append(param_detail)

        return {'type': 'simple', 'params': expected_params}

    def _parse_single_param_expr(self, param_expr: str) -> Optional[Dict[str, Any]]:
        """解析单个参数表达式（如"param>value"）"""
        for op in self.OPERATORS:
            op_index = param_expr.find(op)
            if op_index != -1:
                param_name = param_expr[:op_index].strip()
                param_value = param_expr[op_index + len(op):].strip()

                if not param_name:  # 无效的参数名
                    return None

                return self._parse_param_detail(param_name, param_value, op)

        return None  # 未找到有效运算符

    def _parse_param_detail(self, param_name: str, param_value: str, operator: str) -> Dict[str, Any]:
        """
        解析参数详细信息，统一处理单值和多值参数
        多值参数判断：包含&和:的组合
        """
        # 判断是否为多值参数
        if '&' in param_value and ':' in param_value:
            switches = {}
            for switch_expr in param_value.split('&'):
                if ':' in switch_expr:
                    switch_name, switch_state = switch_expr.split(':', 1)
                    switches[switch_name.strip()] = switch_state.strip()

            return {
                'param_name': param_name,
                'param_type': 'multiple',
                'operator': operator,
                'expected_switches': switches,
                'expected_value': param_value
            }
        else:
            # 单值参数
            return {
                'param_name': param_name,
                'param_type': 'single',
                'operator': operator,
                'expected_value': param_value
            }



    def parse_expected_expression(self, expression: str) -> Dict[str, Any]:
        """
        解析期望值表达式的入口方法

        支持格式:
        1. 简单格式:
           - 参数名1=值1,参数名2>值2,参数名3<=值3
           - 参数名1=k1:开&k2:关&k3:开
        2. 复杂格式:
           - （参数名1=值1 and 参数名2=值2）or（参数名3>值3 and 参数名2!=值2）

        Args:
            expression: 待解析的表达式字符串

        Returns:
            解析结果字典，包含类型和参数信息
        """
        if not expression or expression == 'nan':
            return {'type': 'simple', 'params': []}

        # 检查是否为复杂表达式
        if self._is_complex_expression(expression):
            return self._parse_complex_expression(expression)
        else:
            return self._parse_simple_expression(expression)





        # 检查是否是复杂表达式（包含括号和逻辑运算符）
        if ('(' in expression and ')' in expression) or 'and' in expression  or 'or' in expression or ' and ' in expression or ' or ' in expression:
            logical_parens = self._find_logical_parentheses(expression)
            temp_expr = expression
            for (left_idx, right_idx) in logical_parens:
                inner_expr = temp_expr[left_idx + 1:right_idx].strip()
                temp_expr = f"{temp_expr[:left_idx]}{str(inner_expr).lower()}{temp_expr[right_idx + 1:]}"

            param_names = []
            operators = ['>=', '<=', '!=', '>', '<', '=']
            working_expr = temp_expr
            # 处理表达式，提取参数名
            while working_expr:
                found = False
                for op in operators:
                    op_index = working_expr.find(op)
                    if op_index != -1:
                        param_name = working_expr[:op_index].strip()
                        if param_name:
                            param_names.append(param_name)
                        # 跳过当前操作符和值，继续查找下一个参数
                        next_and = working_expr.find('and', op_index)
                        if next_and != -1:
                            next_and = next_and + 3
                        next_or = working_expr.find('or', op_index)
                        if next_or != -1:
                            next_or = next_or + 2
                        next_delimiter = min(
                            [pos for pos in [next_and, next_or] if pos != -1] or [len(working_expr)])
                        working_expr = working_expr[next_delimiter:].strip()
                        found = True
                        break
                if not found:
                    break
            return {'type': 'complex', 'expression': expression, 'params': param_names}

        # 简单格式：按逗号分割多个参数
        expected_params = []
        param_expressions = [expr.strip() for expr in expression.split(',') if expr.strip()]

        for param_expr in param_expressions:
            if '=' in param_expr:
                param_name, param_value = param_expr.split('=', 1)
                param_name = param_name.strip()
                param_value = param_value.strip()

                # 检查是否是多值参数（包含开关组合）
                if '&' in param_value and ':' in param_value:
                    # 多值参数
                    switches = {}
                    for switch_expr in param_value.split('&'):
                        if ':' in switch_expr:
                            switch_name, switch_state = switch_expr.split(':', 1)
                            switches[switch_name.strip()] = switch_state.strip()

                    expected_params.append({
                        'param_name': param_name,
                        'param_type': 'multiple',
                        'expected_switches': switches,
                        'expected_value': param_value
                    })
                else:
                    # 单值参数
                    expected_params.append({
                        'param_name': param_name,
                        'param_type': 'single',
                        'expected_value': param_value
                    })

        return {'type': 'simple', 'params': expected_params}

    # def validate_complex_expected_expression(self, expression: str, data_row: Dict[str, Any]) -> bool:
    #     """
    #     验证复杂的期望值表达式
    #     支持格式: （参数名1=值1and参数名2=值2）or（参数名3>值3and参数名2!=值2）
    #     """
    #     try:
    #         # 重用条件表达式解析的逻辑
    #         normalized_expr = self._normalize_condition_expression(expression)
    #
    #         def evaluate_single_validation(cond: str) -> bool:
    #             """评估单个验证条件"""
    #             cond = cond.strip()
    #
    #             # 支持的运算符
    #             operators = ['>=', '<=', '!=', '>', '<', '=']
    #
    #             for op in operators:
    #                 if op in cond:
    #                     parts = cond.split(op, 1)
    #                     if len(parts) == 2:
    #                         param_name = parts[0].strip()
    #                         expected_value = parts[1].strip()
    #
    #                         # 获取实际值
    #                         actual_value = str(data_row.get(param_name, '')).strip()
    #
    #                         # 对于多值参数，需要特殊处理
    #                         if '&' in expected_value and ':' in expected_value:
    #                             # 多值参数验证
    #                             return self._validate_multi_value_parameter(actual_value, expected_value)
    #                         else:
    #                             # 单值参数验证
    #                             return self._compare_values(actual_value, op, expected_value)
    #
    #             return False
    #
    #         # 递归处理括号和逻辑运算符
    #         def evaluate_expression(expr: str) -> bool:
    #             expr = expr.strip()
    #
    #             # 处理最内层括号
    #             while '(' in expr and ')' in expr:
    #                 start = expr.rfind('(')
    #                 end = expr.find(')', start)
    #                 if end == -1:
    #                     break
    #
    #                 # 评估括号内的表达式
    #                 inner_expr = expr[start + 1:end]
    #                 result = self._evaluate_simple_expression(inner_expr, evaluate_single_validation)
    #
    #                 # 替换括号及其内容为结果
    #                 expr = expr[:start] + str(result).lower() + expr[end + 1:]
    #
    #             # 评估剩余表达式
    #             return self._evaluate_simple_expression(expr, evaluate_single_validation)
    #
    #         return evaluate_expression(normalized_expr)
    #
    #     except Exception as e:
    #         logger.error(f"验证复杂期望值表达式失败: {expression}, 错误: {str(e)}")
    #         return False
    def _find_logical_parentheses(self,expr: str) -> list[tuple[int, int]]:
        """找逻辑（），（）外面有and和or就是逻辑（）"""
        """识别表达式中用于分组的逻辑括号，排除参数中的字面括号"""
        expr = expr.strip()
        logical_parens = []
        left_paren_positions = []

        for idx, char in enumerate(expr):
            if char == '(':
                # 检查左括号是否为逻辑括号
                is_logical_left = False
                if idx == 0:
                    is_logical_left = True
                else:
                    prev_non_space = idx - 1
                    while prev_non_space >= 0 and expr[prev_non_space].isspace():
                        prev_non_space -= 1
                    if prev_non_space >= 0:
                        if (prev_non_space >= 2 and expr[prev_non_space - 2:prev_non_space + 1] == "and") or \
                                (prev_non_space >= 1 and expr[prev_non_space - 1:prev_non_space + 1] == "or"):
                            is_logical_left = True
                if is_logical_left:
                    left_paren_positions.append(idx)

            elif char == ')':
                if not left_paren_positions:
                    continue
                left_idx = left_paren_positions.pop()
                # 检查右括号是否为逻辑括号
                is_logical_right = False
                if idx == len(expr) - 1:
                    is_logical_right = True
                else:
                    next_non_space = idx + 1
                    while next_non_space < len(expr) and expr[next_non_space].isspace():
                        next_non_space += 1
                    if next_non_space < len(expr):
                        if (next_non_space + 2 < len(expr) and expr[next_non_space:next_non_space + 3] == "and") or \
                                (next_non_space + 1 < len(expr) and expr[
                                    next_non_space:next_non_space + 2] == "or"):
                            is_logical_right = True
                if is_logical_right:
                    logical_parens.append((left_idx, idx))

        return sorted(logical_parens, key=lambda x: x[0], reverse=True)

    def validate_complex_expected_expression(self,expression: str, data_row: Dict[str, Any]) -> bool:
        """
        验证复杂的期望值表达式（在原有基础上修改）
        支持参数名/值包含括号的情况
        """



        try:
            normalized_expr = self._normalize_condition_expression(expression)
            logical_parens = self._find_logical_parentheses(normalized_expr)

            def evaluate_single_validation(cond: str) -> bool:
                """评估单个验证条件（保持原有逻辑结构）"""
                cond = cond.strip()
                if not cond:
                    return False

                operators = ['>=', '<=', '!=', '>', '<', '=']
                op_found = None
                op_index = -1

                for op in operators:
                    op_index = cond.find(op)
                    if op_index != -1:
                        op_found = op
                        break
                if op_found is None:
                    return False

                param_name = cond[:op_index].strip()
                expected_value = cond[op_index + len(op_found):].strip()
                actual_value = str(data_row.get(param_name, '')).strip()

                if '&' in expected_value and ':' in expected_value:
                    return self._validate_multi_value_parameter(actual_value, expected_value)
                else:
                    return self._compare_values(actual_value, op_found, expected_value)

            def evaluate_expression(expr: str) -> bool:
                expr = expr.strip()
                if not expr:
                    return False

                # 使用新的逻辑括号处理方式替换原有盲找括号的逻辑
                for (left_idx, right_idx) in logical_parens:
                    inner_expr = expr[left_idx + 1:right_idx].strip()
                    if not inner_expr:
                        continue
                    inner_result = self._evaluate_simple_expression(inner_expr, evaluate_single_validation)
                    expr = f"{expr[:left_idx]}{str(inner_result).lower()}{expr[right_idx + 1:]}"

                return self._evaluate_simple_expression(expr, evaluate_single_validation)

            return evaluate_expression(normalized_expr)

        except Exception as e:
            logger.error(f"验证复杂期望值表达式失败: {expression}, 错误: {str(e)}")
            return False

    def _validate_multi_value_parameter(self, actual_value: str, expected_value: str) -> bool:
        """验证多值参数"""
        if not actual_value or not expected_value:
            return False

        # 解析期望的开关状态
        expected_switches = {}
        for switch_expr in expected_value.split('&'):
            if ':' in switch_expr:
                switch_name, switch_state = switch_expr.split(':', 1)
                expected_switches[switch_name.strip()] = switch_state.strip()

        # 使用现有的多值匹配逻辑
        is_match, _ = self._check_multi_value_match(actual_value, expected_switches)
        return is_match

    def execute_validation_rule(self, rule_id: str, data_groups: Dict[str, pd.DataFrame],sector_id) -> List[Dict[str, Any]]:
        """执行单个验证规则"""
        if rule_id not in self.validation_rules:
            logger.warning(f"验证规则 {rule_id} 不存在")
            return []

        rule = self.validation_rules[rule_id]
        errors = []

        logger.info(f"执行验证规则: {rule_id} ({rule['check_type']})")

        # 根据校验类型执行不同的验证
        if rule['check_type'] == '漏配':
            errors.extend(self._check_missing_config(rule, data_groups,sector_id))
        elif rule['check_type'] == '错配':
            errors.extend(self._check_incorrect_config(rule, data_groups,sector_id))
        else:
            logger.warning(f"未知的校验类型: {rule['check_type']}")

        # 如果当前规则通过且有继续校验，执行继续校验
        if not errors and rule['next_check_id']:
            logger.info(f"规则 {rule_id} 通过，继续执行: {rule['next_check_id']}")
            errors.extend(self.execute_validation_rule(rule['next_check_id'], data_groups,sector_id))
        elif errors:
            logger.info(f"规则 {rule_id} 检查失败，不继续后续验证")

        return errors

    def _check_missing_config(self, rule: Dict[str, Any], data_groups: Dict[str, pd.DataFrame],sector_id) -> List[Dict[str, Any]]:
        """检查漏配"""
        mo_name = rule['mo_name']
        condition_expr = rule['condition_expression']
        expected_expr = rule['expected_expression']

        errors = []

        if mo_name not in data_groups:
            errors.append({
                'sector_id': "",
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'check_type': '漏配',
                'error_type': '数据不存在',
                'message': f'{mo_name}数据不存在',
                'error_description': rule['error_description']
            })
            return errors

        mo_data = pd.DataFrame(data_groups[mo_name])
        expected_result = self.parse_expected_expression(expected_expr)

        if expected_result['type'] == 'simple' and not expected_result['params']:
            logger.warning(f"规则 {rule['rule_id']} 没有有效的期望值表达式")
            return errors


        # 检查是否存在符合条件的记录
        found_matching_record = False

        for _, row in mo_data.iterrows():
            row_dict = row.to_dict()
            row_dict = {k: str(v).strip() for k, v in row_dict.items()}

            # 检查条件表达式
            if not self.parse_condition_expression(condition_expr, row_dict):
                continue

            # 检查期望值
            if expected_result['type'] == 'complex':
                # 复杂表达式验证
                if self.validate_complex_expected_expression(expected_result['expression'], row_dict):
                    found_matching_record = True
                    break
            else:
                # 简单表达式验证
                all_params_match = True
                for expected_param in expected_result['params']:
                    param_name = expected_param['param_name']

                    if param_name not in row_dict:
                        all_params_match = False
                        break

                    if expected_param['param_type'] == 'multiple':
                        # 多值参数检查
                        actual_value = row_dict[param_name]
                        is_match, wrong_switches = self._check_multi_value_match(actual_value,
                                                                                 expected_param['expected_switches'])
                        if not is_match:
                            all_params_match = False
                            break
                    else:
                        # 单值参数检查
                        if row_dict[param_name] != expected_param['expected_value']:
                            all_params_match = False
                            break

                if all_params_match:
                    found_matching_record = True
                    break

        if not found_matching_record:
            # 获取参数名称列表
            if expected_result['type'] == 'complex':
                import re
                param_names = re.findall(r'([^=<>!]+)=', expected_result['expression'])
                param_names = [name.strip() for name in param_names]
            else:
                param_names = [p['param_name'] for p in expected_result['params']]

            errors.append({
                'sector_id': sector_id,
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'param_names': param_names,
                'check_type': '漏配',
                'error_type': '漏配',
                'message': f'未找到符合条件的配置记录',
                'condition': condition_expr,
                'expected_expression': expected_expr,
                'error_description': rule['error_description']
            })

        return errors

    def _check_incorrect_config(self, rule: Dict[str, Any], data_groups: Dict[str, pd.DataFrame],sector_id) -> List[Dict[str, Any]]:
        """检查错配"""
        mo_name = rule['mo_name']
        condition_expr = rule['condition_expression']
        expected_expr = rule['expected_expression']

        errors = []

        if mo_name not in data_groups:
            errors.append({
                'sector_id': "",
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'check_type': '错配',
                'error_type': '数据不存在',
                'message': f'{mo_name}数据不存在',
                'error_description': rule['error_description']
            })
            return errors  # 数据不存在时不报错配错误

        mo_data = pd.DataFrame(data_groups[mo_name])
        expected_result = self.parse_expected_expression(expected_expr)

        if expected_result['type'] == 'simple' and not expected_result['params']:
            logger.warning(f"规则 {rule['rule_id']} 没有有效的期望值表达式")
            return errors

        # 检查每条记录
        for idx, row in mo_data.iterrows():
            row_dict = row.to_dict()
            row_dict = {k: str(v).strip() for k, v in row_dict.items()}

            # 检查条件表达式
            if not self.parse_condition_expression(condition_expr, row_dict):
                continue

            # 检查期望值
            if expected_result['type'] == 'complex':
                # 复杂表达式验证
                if not self.validate_complex_expected_expression(expected_result['expression'], row_dict):
                    # 复杂表达式验证失败，生成错误
                    for expected_param in expected_result['params']:
                        param_name = expected_param['param_name']

                        if param_name not in row_dict:
                            continue

                        if expected_param['param_type'] == 'multiple':
                            # 多值参数检查
                            actual_value = row_dict[param_name]
                            is_match, wrong_switches = self._check_multi_value_match(actual_value,
                                                                                     expected_param[
                                                                                         'expected_switches'])
                            if not is_match:
                                # 获取参数的值描述，解析每个开关的说明
                                value_description = self._get_parameter_value_description(mo_name, param_name)
                                switch_descriptions = self._parse_value_descriptions(value_description)

                                # 只为错误的开关生成错误描述
                                error_switch_descriptions = []
                                for wrong_switch in wrong_switches:
                                    switch_name = wrong_switch['switch_name']
                                    if switch_name in switch_descriptions:
                                        error_switch_descriptions.append(
                                            f"{switch_name}: {switch_descriptions[switch_name]}")

                                errors.append({
                                    'sector_id': row_dict.get('f_site_id', "") + "_" + row_dict.get('f_cell_id', ""),
                                    'rule_id': rule['rule_id'],
                                    'mo_name': mo_name,
                                    'param_name': param_name,
                                    'check_type': '错配',
                                    'error_type': '错配',
                                    'message': f'{param_name}开关配置错误',
                                    'current_value': actual_value,
                                    'expected_value': expected_param['expected_value'],
                                    'wrong_switches': wrong_switches,
                                    'switch_descriptions': error_switch_descriptions,
                                    'condition': condition_expr,
                                    'error_description': rule['error_description'],
                                    'row_index': idx
                                })
                        else:
                            # 单值参数检查
                            if row_dict[param_name] != expected_param['expected_value']:
                                errors.append({
                                    'sector_id': row_dict.get('f_site_id', "") + "_" + row_dict.get('f_cell_id', ""),
                                    'rule_id': rule['rule_id'],
                                    'mo_name': mo_name,
                                    'param_name': param_name,
                                    'check_type': '错配',
                                    'error_type': '错配',
                                    'message': f'{param_name}配置错误',
                                    'current_value': row_dict[param_name],
                                    'expected_value': expected_param['expected_value'],
                                    'condition': condition_expr,
                                    'error_description': rule['error_description'],
                                    'row_index': idx
                                })
            else:
                # 简单表达式验证
                for expected_param in expected_result['params']:
                    param_name = expected_param['param_name']

                    if param_name not in row_dict:
                        continue

                    if expected_param['param_type'] == 'multiple':
                        # 多值参数检查
                        actual_value = row_dict[param_name]
                        is_match, wrong_switches = self._check_multi_value_match(actual_value,
                                                                                 expected_param['expected_switches'])
                        if not is_match:
                            # 获取参数的值描述，解析每个开关的说明
                            value_description = self._get_parameter_value_description(mo_name, param_name)
                            switch_descriptions = self._parse_value_descriptions(value_description)

                            # 只为错误的开关生成错误描述
                            error_switch_descriptions = []
                            for wrong_switch in wrong_switches:
                                switch_name = wrong_switch['switch_name']
                                if switch_name in switch_descriptions:
                                    error_switch_descriptions.append(
                                        f"{switch_name}: {switch_descriptions[switch_name]}")

                            errors.append({
                                'sector_id': row_dict.get('f_site_id',"")+"_"+row_dict.get('f_cell_id',""),
                                'rule_id': rule['rule_id'],
                                'mo_name': mo_name,
                                'param_name': param_name,
                                'check_type': '错配',
                                'error_type': '错配',
                                'message': f'{param_name}开关配置错误',
                                'current_value': actual_value,
                                'expected_value': expected_param['expected_value'],
                                'wrong_switches': wrong_switches,
                                'switch_descriptions': error_switch_descriptions,
                                'condition': condition_expr,
                                'error_description': rule['error_description'],
                                'row_index': idx
                            })
                    else:
                        # 单值参数检查
                        if row_dict[param_name] != expected_param['expected_value']:
                            errors.append({
                                'sector_id': row_dict.get('f_site_id',"")+"_"+row_dict.get('f_cell_id',""),
                                'rule_id': rule['rule_id'],
                                'mo_name': mo_name,
                                'param_name': param_name,
                                'check_type': '错配',
                                'error_type': '错配',
                                'message': f'{param_name}配置错误',
                                'current_value': row_dict[param_name],
                                'expected_value': expected_param['expected_value'],
                                'condition': condition_expr,
                                'error_description': rule['error_description'],
                                'row_index': idx
                            })

        return errors

    def _check_multi_value_match(self, actual_value: str, expected_switches: Dict[str, str]) -> Tuple[bool, List[Dict[str, str]]]:
        """
        检查多值参数是否匹配
        返回: (是否匹配, 错误的开关列表)
        """
        if not actual_value or not expected_switches:
            return False, []

        # 解析实际值中的开关状态
        actual_switches = {}
        for switch_expr in actual_value.split('&'):
            if ':' in switch_expr:
                switch_name, switch_state = switch_expr.split(':', 1)
                actual_switches[switch_name.strip()] = switch_state.strip()

        # 检查每个期望的开关状态，收集错误的开关
        wrong_switches = []
        all_match = True

        for switch_name, expected_state in expected_switches.items():
            if switch_name not in actual_switches:
                wrong_switches.append({
                    'switch_name': switch_name,
                    'expected_state': expected_state,
                    'actual_state': '未配置',
                    'error_type': '缺失'
                })
                all_match = False
            elif actual_switches[switch_name] != expected_state:
                wrong_switches.append({
                    'switch_name': switch_name,
                    'expected_state': expected_state,
                    'actual_state': actual_switches[switch_name],
                    'error_type': '错误'
                })
                all_match = False

        return all_match, wrong_switches

    def _get_parameter_value_description(self, mo_name: str, param_name: str) -> str:
        """获取参数的值描述"""
        if mo_name in self.parameter_info and param_name in self.parameter_info[mo_name]['parameters']:
            return self.parameter_info[mo_name]['parameters'][param_name].get('value_description', '')
        return ''

    def _parse_value_descriptions(self, value_description: str) -> Dict[str, str]:
        """
        解析值描述字符串，提取各个开关的说明
        格式: "beam1:第一波束开关,beam2:第二波束开关,beam3:第三波束开关"
        """
        descriptions = {}
        if not value_description:
            return descriptions

        for desc_part in value_description.split(','):
            if ':' in desc_part:
                switch_name, switch_desc = desc_part.split(':', 1)
                descriptions[switch_name.strip()] = switch_desc.strip()

        return descriptions

    def validate_sector_data(self, data_groups: Dict[str, pd.DataFrame],sector_id) -> List[Dict[str, Any]]:
        """验证扇区数据"""
        all_errors = []

        # 找到所有入口验证规则（没有被其他规则引用的规则）
        referenced_rules = set()
        for rule in self.validation_rules.values():
            if rule['next_check_id']:
                referenced_rules.add(rule['next_check_id'])

        entry_rules = [rule_id for rule_id in self.validation_rules.keys()
                       if rule_id not in referenced_rules]

        logger.info(f"发现 {len(entry_rules)} 个入口验证规则: {entry_rules}")

        # 执行每个入口规则
        for rule_id in entry_rules:
            errors = self.execute_validation_rule(rule_id, data_groups,sector_id)
            all_errors.extend(errors)

        return all_errors

    def create_sample_excel(self) -> None:
        """创建示例Excel文件"""
        logger.info("正在生成示例参数知识库...")

        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 创建参数信息表
        self._create_parameter_info_sheet(wb)

        # 创建验证规则表
        self._create_validation_rules_sheet(wb)

        # 保存文件
        wb.save(self.knowledge_file)
        logger.info(f"示例参数知识库已生成: {self.knowledge_file}")

    def _create_parameter_info_sheet(self, wb: Workbook) -> None:
        """创建参数信息表"""
        ws = wb.create_sheet("参数信息")

        # 设置表头
        headers = ['MO名称', 'MO描述', '场景类型', '参数名称', '参数ID', '参数类型', '参数含义', '值描述']
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 添加示例数据
        sample_data = [
            # NRCELL参数
            ["NRCELL", "5G小区对象", "5G基础配置", "跟踪区码", "tac", "single", "标识小区所属的跟踪区", ""],
            ["NRCELL", "5G小区对象", "5G基础配置", "小区状态", "cellState", "single", "小区的激活状态", ""],

            # NRDUCELL参数
            ["NRDUCELL", "5G DU小区对象", "5G物理层配置", "小区半径(米)", "cellRadius", "single", "小区覆盖半径", ""],
            ["NRDUCELL", "5G DU小区对象", "5G物理层配置", "最大传输功率", "maxTxPower", "single", "小区最大发射功率",
             ""],

            # NRDUCELLBEAM参数（多值参数）
            ["NRDUCELLBEAM", "5G波束配置对象", "5G波束管理", "波束开关组合", "beamSwitchComb", "multiple",
             "波束开关状态组合", "beam1:第一波束开关,beam2:第二波束开关,beam3:第三波束开关"],

            # NRCELLFREQRELATION参数
            ["NRCELLFREQRELATION", "小区频率关系对象", "频率管理配置", "连接态频率优先级", "connectedFreqPriority",
             "single", "连接态下的频率优先级", ""],
        ]

        for row_data in sample_data:
            ws.append(row_data)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_validation_rules_sheet(self, wb: Workbook) -> None:
        """创建验证规则表"""
        ws = wb.create_sheet("验证规则")

        # 设置表头
        headers = ['校验ID', '校验类型', 'MO名称', '条件表达式', '期望值表达式', '错误描述', '继续校验ID']
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 添加示例验证规则
        sample_rules = [
            # 复杂嵌套验证链示例
            ["MISS_001", "漏配", "NRCELL", "", "跟踪区码=100", "缺少跟踪区码为100的小区配置", "ERROR_001"],
            ["ERROR_001", "错配", "NRDUCELL", "跟踪区码=100", "小区半径(米)=500",
             "跟踪区码为100的小区，半径应配置为500米", "ERROR_002"],
            ["ERROR_002", "错配", "NRDUCELL", "跟踪区码=100and小区半径(米)=500", "最大传输功率=43",
             "半径500米的小区，功率应为43dBm", "ERROR_003"],
            ["ERROR_003", "错配", "NRDUCELLBEAM", "跟踪区码=100", "波束开关组合=beam1:开&beam2:关&beam3:开",
             "跟踪区码100的小区，波束组合应为beam1开beam2关beam3开", "MISS_002"],
            ["MISS_002", "漏配", "NRCELLFREQRELATION", "跟踪区码=100", "连接态频率优先级=1",
             "缺少跟踪区码100小区的频率优先级配置", "ERROR_004"],
            ["ERROR_004", "错配", "NRCELL", "跟踪区码=100and连接态频率优先级=1", "小区状态=激活",
             "已配置频率优先级的小区状态应为激活", ""],

            # 复杂条件示例
            ["COMPLEX_001", "错配", "NRDUCELL", "(跟踪区码=200or跟踪区码=300)and小区状态=激活", "小区半径(米)=1000",
             "特殊跟踪区的激活小区半径应为1000米", ""],
            ["COMPLEX_002", "漏配", "NRCELL", "小区半径(米)>500and最大传输功率>=40", "小区状态=激活",
             "大半径高功率小区必须激活", ""],
        ]

        for row_data in sample_rules:
            ws.append(row_data)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def run_validation_example(self) -> None:
        """运行验证示例"""
        logger.info("🧪 开始验证示例测试...")

        # 创建测试数据
        # 创建测试数据
        datas = {
            "result": {
                "NRDUCELL": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR DU小区标识": "4",
                        "小区半径(米)": "4000"
                    }
                ],
                "NRCELLALGOSWITCH": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "异频切换算法开关": "基于覆盖的异频切换开关:开&基于频率优先级的异频切换开关:关&异频重定向开关:开&基于运营商专用优先级的异频切换开关:关&音视频异频切换配合开关:开&基于业务的异频切换开关:关&基于覆盖的异频盲切换开关:关&FR1到FR2频点轮询选择开关:关&基于上行干扰的异频切换开关:关&基于SSB SINR的异频切换开关:关&NSA基于上行干扰的异频切换开关:关&异频切换配合开关:关&基于能效的异频切换开关:关&基于MBS兴趣指示的异频切换开关:关&基于业务的异频盲切换开关:关"
                    }
                ],
                "NRCELLFREQRELATION": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7783",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "14"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "5361",
                        "连接态频率优先级": "1",
                        "小区重选优先级": "5",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7714",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7853",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    }
                ],
                "NRCELLINTERFHOMEAGRP": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "异频切换测量参数组标识": "0",
                        "基于覆盖的异频A5 RSRP触发门限1(dBm)": "-105",
                        "基于覆盖的异频A5 RSRP触发门限2(dBm)": "-100",
                        "基于覆盖的异频A2 RSRP触发门限(dBm)": "-105",
                        "基于覆盖的异频A1 RSRP触发门限(dBm)": "-100",
                        "异频测量事件时间迟滞(毫秒)": "320",
                        "异频测量事件幅度迟滞(0.5dB)": "2",
                        "异频A1A2时间迟滞(毫秒)": "320",
                        "异频A1A2幅度迟滞(0.5dB)": "2"
                    }
                ],
                "NRCELLQCIBEARER": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "1",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "2",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "3",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "4",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "5",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "6",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "7",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "8",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "9",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "65",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "66",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "69",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "70",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "75",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "79",
                        "异频切换测量参数组标识": "0"
                    }
                ],
                "NRCELLRESELCONFIG": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "非同频测量RSRP触发门限(2dB)": "10",
                        "服务频点低优先级RSRP重选门限(2dB)": "9"
                    }
                ]
            }
        }
        test_datas = datas['result']
        # 执行验证
        sector_datas = {}
        for mo, raw_data in test_datas.items():
            data_df = pd.DataFrame(raw_data)

            # 按站点和小区ID分组
            for (site_id, cell_id), sector_df in data_df.groupby(
                    ['f_site_id', 'f_cell_id'],
                    dropna=False
            ):
                sector_id = f"{site_id}_{cell_id}"

                # 如果sector_id不在外层字典中，初始化一个空字典
                if sector_id not in sector_datas:
                    sector_datas[sector_id] = {}

                # 将当前MO的数据放入对应的sector_id下
                sector_datas[sector_id][mo] = sector_df


        for sector_id,sector_dfs in sector_datas.items():
            errors = self.validate_sector_data(sector_dfs,sector_id)

        # 输出结果
        if errors:
            logger.info(f"🔍 发现 {len(errors)} 个验证问题:")
            for i, error in enumerate(errors, 1):
                logger.info(f"   {i}. 【{error['check_type']}】{error.get('rule_id', 'N/A')} - {error['mo_name']}")
                if 'param_name' in error:
                    logger.info(f"      参数: {error['param_name']}")
                if 'param_names' in error:
                    logger.info(f"      参数: {', '.join(error['param_names'])}")
                logger.info(f"      错误: {error['message']}")

                # 处理多值参数的开关错误详情
                if 'wrong_switches' in error and error['wrong_switches']:
                    logger.info(f"      开关错误详情:")
                    for switch_error in error['wrong_switches']:
                        logger.info(
                            f"        - {switch_error['switch_name']}: 期望{switch_error['expected_state']}, 实际{switch_error['actual_state']}")

                    # 显示错误开关的描述
                    if 'switch_descriptions' in error and error['switch_descriptions']:
                        logger.info(f"      开关说明:")
                        for desc in error['switch_descriptions']:
                            logger.info(f"        - {desc}")

                # 显示单值参数的期望值和实际值
                elif 'current_value' in error and 'expected_value' in error:
                    logger.info(f"      期望值: {error['expected_value']}")
                    logger.info(f"      实际值: {error['current_value']}")

                if error.get('error_description'):
                    logger.info(f"      说明: {error['error_description']}")
                logger.info("")
        else:
            logger.info("✅ 所有验证规则都通过了")


def main():
    """主程序入口"""
    try:
        logger.info("🚀 启动参数核查系统...")

        # 创建参数核查器实例
        checker = ParameterChecker()

        # 运行验证示例
        checker.run_validation_example()

        logger.info("✨ 参数核查系统运行完成！")
        logger.info("📋 系统特性:")
        logger.info("   • 双分表设计：参数信息与验证规则完全分离")
        logger.info("   • 复杂条件支持：(param1=value1and param2=value2)or(param3>value3)")
        logger.info("   • 嵌套验证链：支持漏配↔错配无限嵌套调用")
        logger.info("   • 多值参数处理：beam1:开&beam2:关&beam3:开格式")
        logger.info("   • 智能条件筛选：先筛选符合条件的行再进行验证")

        return True

    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}")
        return False


if __name__ == "__main__":
    success = main()
    if not success:
        exit(1)
